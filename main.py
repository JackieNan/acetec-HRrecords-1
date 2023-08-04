import pandas as pd
import openpyxl
import gui
import os
from openpyxl.styles import PatternFill
from openpyxl.utils import datetime
from datetime import datetime
mainpath = os.getcwd()
def main():

    gui.run_gui()
    df = pd.read_excel('data1.xlsx', sheet_name='考勤汇总')
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    #将异常数据用红色填充
    wb = openpyxl.load_workbook('data1.xlsx', data_only=True)
    ws = wb['考勤汇总']
    for i in range(2, len(df) + 2):
        if df.iloc[i - 2, 12] > 60:
            cell = 'M' + str(i)
            ws[cell].fill = red_fill
    #将data2工号列后四位保存到列表，s2 = df2工号
    workbook = openpyxl.load_workbook('data2.xlsx', data_only=True)
    worksheet = workbook['请假']
    worksheet_1 = workbook['异常']

    s2 = []
    for i in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=i, column=25).value
        #if cell_value is not None:
        s2.append(str(cell_value)[-4:])



    #将data2日期保存到列表，s3 = 休假日期，s4 = 截止日期
    s3 = []
    for i in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=i, column=32).value
        s3.append(str(cell_value))

    for k in range(0, len(s3)):
        if s3[k] is None or s3[k] == 'None':
            # 如果字符串为空，则跳过本次循环
            continue
        try:
            # 尝试使用第一种日期格式解析字符串
            s3[k] = datetime.strptime(s3[k], '%Y/%m/%d')
        except ValueError:
            # 如果第一种日期格式解析失败，则尝试使用第二种日期格式
            s3[k] = datetime.strptime(s3[k], '%Y-%m-%d %H:%M:%S')
    s4 = []
    for i in range(2, worksheet.max_row + 1):
        cell_value = worksheet.cell(row=i, column=33).value
        s4.append(str(cell_value))
    for k in range(0, len(s4)):
        if s4[k] is None or s4[k] == 'None':
            # 如果字符串为空，则跳过本次循环
            continue
        try:
            # 尝试使用第一种日期格式解析字符串
            s4[k] = datetime.strptime(s4[k], '%Y/%m/%d')
        except ValueError:
            # 如果第一种日期格式解析失败，则尝试使用第二种日期格式
            s4[k] = datetime.strptime(s4[k], '%Y-%m-%d %H:%M:%S')
    s2_1 = []
    for i in range(2, worksheet_1.max_row + 1):
        cell_value = worksheet_1.cell(row=i, column=43).value
        #if cell_value is not None:
        s2_1.append(str(cell_value)[-4:])

    s3_1 = []
    for i in range(2, worksheet_1.max_row + 1):
        cell_value = worksheet_1.cell(row=i, column=45).value
        s3_1.append(str(cell_value))

    for k in range(0, len(s3_1)):
        if s3_1[k] == 'na':
            # 如果字符串为空，则跳过本次循环
            continue
        try:
            # 尝试使用第一种日期格式解析字符串
            s3_1[k] = datetime.strptime(s3_1[k], '%Y/%m/%d')
        except ValueError:
            # 如果第一种日期格式解析失败，则尝试使用第二种日期格式
            s3_1[k] = datetime.strptime(s3_1[k], '%Y-%m-%d %H:%M:%S')

    #匹配
    for i in range(2, ws.max_row + 1):
        for j in range(5, 14):
            if ws.cell(row=i, column=j).fill.start_color.index in ['FFFF0000', 'FFFFA500']:
                b_value = ws.cell(row=i, column=2).value
                d_value = ws.cell(row=i, column=4).value
                c_value = ws.cell(row=i, column=8).value

                if (type(d_value) != datetime):
                    try:
                        # 尝试使用第一种日期格式解析字符串
                        d_value = datetime.strptime(d_value, '%Y/%m/%d')

                    except ValueError:
                        # 如果第一种日期格式解析失败，则尝试使用第二种日期格式
                        d_value = datetime.strptime(d_value, '%Y-%m-%d %H:%M:%S')

                for k in range(0, len(s3)):
                    if (s4[k] is None or s4[k] == 'None') or (s3[k] == 'None'or s3[k] is None):
                        # 如果字符串为空，则跳过本次循环
                        continue

                    if (s2[k] == b_value) and (s3[k] <= d_value <= s4[k]):
                        w_value = s2[k]
                        z_value = worksheet.cell(row=k+2, column=26).value
                        af_value = worksheet.cell(row=k+2, column=32).value
                        ai_value = worksheet.cell(row=k+2, column=35).value
                        ag_value = worksheet.cell(row=k+2, column=33).value
                        aj_value = worksheet.cell(row=k+2, column=36).value
                        ws.cell(row=i, column=14).value = f'{w_value},{z_value},{af_value},{ai_value},{ag_value},{aj_value}'
                        break
                    else:
                        ws.cell(row=i, column=14).value = '无oa记录'
                        continue

                if ((c_value == '夜') and not (
                        ws.cell(row=i, column=5).fill.start_color.index in ['FFFF0000', 'FFFFA500'])
                        and (ws.cell(row=i, column=6).fill.start_color.index in ['FFFF0000'])):
                    ws.cell(row=i,column=15).value = '夜班下班忘记打卡'
                    break
                for k in range(0, len(s3_1)):
                    if s3_1 == 'None':
                        continue

                    if ((s2_1[k] == b_value) and (s3_1[k] == d_value)):
                        w_value = s2_1[k]
                        z_value = worksheet_1.cell(row=k + 2, column=42).value
                        af_value = worksheet_1.cell(row=k + 2, column=45).value
                        ai_value = worksheet_1.cell(row=k + 2, column=46).value
                        ag_value = worksheet_1.cell(row=k + 2, column=47).value
                        ws.cell(row=i,
                                column=15).value = f'{w_value},{z_value},{af_value},{ai_value},{ag_value}'
                        break
                    else:
                        ws.cell(row=i, column=15).value = '无异常信息'
                        continue

    wb.save('结果.xlsx')
if __name__ == '__main__':
    main()